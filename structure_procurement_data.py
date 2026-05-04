"""
Procurement Data Structure & Import Tool
==========================================
Converts raw quotation data into structured procurement format
with proper columns for analysis and tracking.
"""

import pandas as pd
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class ProcurementDataStructurer:
    """Converts raw quotation data into structured format"""
    
    # Define the standard procurement column structure
    PROCUREMENT_COLUMNS = [
        'Date',
        'Item Description',
        'Category',
        'Quantity',
        'Unit',
        'Unit Price',
        'Amount',
        'Budget',
        'Status',
        'Vendor',
        'Payment Method',
        'Notes'
    ]
    
    def __init__(self, input_file: str):
        self.input_file = input_file
        self.structured_data = []
        
    def parse_quotation(self):
        """Parse the quotation Excel file and extract structured data"""
        
        # Read raw data
        df_raw = pd.read_excel(self.input_file, sheet_name=0, header=None)
        
        # Extract metadata from header rows
        project_date = "22/04/2026"  # From row 5
        project_location = "Agbowa Ikorodu"
        company = "EXIMPS&CLOVES INFRASTRUCTURE LIMITED"
        
        items = []
        
        # Parse FENCING QUOTATION (rows 12-23, 0-indexed: 13-24)
        fencing_rows = [
            ("Cement", 300, "Bags", 13000, 3900000),
            ("Blocks", 15000, "Pcs", 600, 9000000),
            ("Water Container Truck", 3, "Trips", 15000, 45000),
            ("Labour (10 Workers)", 1, "Lot", 1000000, 1000000),
            ("Sharp Sand", 3, "Tons", 3000000, 900000),
            ("Gravel", 3, "Tons", 350000, 1050000),
            ("Iron Rods (12mm)", 50, "Rods", 10000, 500000),
            ("Binding Wire", 5, "Coils", 7000, 35000),
            ("Transportation", 1, "Lot", 100000, 100000),
            ("Accommodation", 1, "Lot", 100000, 100000),
            ("Miscellaneous", 1, "Lot", 100000, 50000)
        ]
        
        for description, qty, unit, unit_price, amount in fencing_rows:
            items.append({
                'Date': project_date,
                'Item Description': description,
                'Category': 'Fencing',
                'Quantity': qty,
                'Unit': unit,
                'Unit Price': unit_price,
                'Amount': amount,
                'Budget': amount,  # Assume budget equals quotation amount
                'Status': 'Pending',  # Initial status
                'Vendor': company,
                'Payment Method': 'TBD',
                'Notes': f'Fencing works for {project_location}'
            })
        
        # Parse SITE CLEARING QUOTATION (rows 32-37, 0-indexed: 33-38)
        clearing_rows = [
            ("Excavator", 3, "Days", 450000, 1350000),
            ("Pay Loader", 3, "Days", 380000, 1140000),
            ("Hotel Accommodation", 3, "Days", 20000, 60000),
            ("Feeding Allowance", 3, "Days", 10000, 30000),
            ("Lowbed Truck", 2, "Days", 400000, 800000)
        ]
        
        for description, qty, unit, unit_price, amount in clearing_rows:
            items.append({
                'Date': project_date,
                'Item Description': description,
                'Category': 'Site Clearing',
                'Quantity': qty,
                'Unit': unit,
                'Unit Price': unit_price,
                'Amount': amount,
                'Budget': amount,
                'Status': 'Pending',
                'Vendor': company,
                'Payment Method': 'TBD',
                'Notes': f'Site clearing for {project_location}'
            })
        
        self.structured_data = items
        print(f"✅ Parsed {len(items)} procurement items")
        return items
    
    def create_structured_file(self, output_file: str = None):
        """Create a new Excel file with structured data"""
        
        if not self.structured_data:
            self.parse_quotation()
        
        if output_file is None:
            output_file = self.input_file.replace('.xlsx', '_STRUCTURED.xlsx')
        
        # Create DataFrame
        df = pd.DataFrame(self.structured_data)
        df = df[self.PROCUREMENT_COLUMNS]  # Ensure column order
        
        # Create Excel file with formatting
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Procurement', index=False)
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Procurement']
            
            # Format header row
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border
            
            # Format data rows
            for row_num, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(df)+1), 2):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    
                    # Format numeric columns
                    if cell.column in [5, 6, 7, 8]:  # Unit Price, Amount, Budget columns
                        cell.number_format = '₦#,##0'
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    
                    # Format quantity column
                    if cell.column == 4:
                        cell.number_format = '0'
                        cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Adjust column widths
            column_widths = {
                'A': 12,  # Date
                'B': 25,  # Item Description
                'C': 15,  # Category
                'D': 12,  # Quantity
                'E': 10,  # Unit
                'F': 15,  # Unit Price
                'G': 15,  # Amount
                'H': 15,  # Budget
                'I': 12,  # Status
                'J': 20,  # Vendor
                'K': 15,  # Payment Method
                'L': 30   # Notes
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
            
            # Freeze header row
            worksheet.freeze_panes = 'A2'
        
        print(f"✅ Structured file created: {output_file}")
        return output_file
    
    def add_metadata_sheet(self, file_path: str):
        """Add metadata sheet with project information"""
        
        from openpyxl.utils import get_column_letter
        
        workbook = openpyxl.load_workbook(file_path)
        
        # Create metadata sheet
        if 'Project Info' in workbook.sheetnames:
            del workbook['Project Info']
        
        ws = workbook.create_sheet('Project Info', 0)
        
        metadata = [
            ['PROCUREMENT PROJECT - PROJECT INFORMATION'],
            [''],
            ['Project Name:', 'Estate Procurement - Agbowa Ikorodu'],
            ['Company:', 'EXIMPS&CLOVES INFRASTRUCTURE LIMITED'],
            ['Project Location:', 'Agbowa Ikorodu, Ogun State'],
            ['Quotation Date:', '22/04/2026'],
            ['Report Generated:', datetime.now().strftime('%d/%m/%Y %H:%M:%S')],
            [''],
            ['FINANCIAL SUMMARY'],
            ['Total Fencing Cost:', '=SUM(\'Procurement\'!G2:G12)'],
            ['Total Site Clearing Cost:', '=SUM(\'Procurement\'!G13:G17)'],
            ['Grand Total:', '=SUM(\'Procurement\'!G:G)'],
            [''],
            ['NOTES:'],
            ['- Status defaults to "Pending" for all items'],
            ['- Payment Method should be updated as per agreement'],
            ['- Budget equals quoted amount initially'],
            ['- Track actual spending in separate columns as needed']
        ]
        
        for row_num, row_data in enumerate(metadata, 1):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                
                if row_num == 1:
                    cell.font = Font(bold=True, size=14, color="FFFFFF")
                    cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                elif row_num == 9:
                    cell.font = Font(bold=True, size=12, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                elif row_num == 14:
                    cell.font = Font(bold=True, size=11)
                elif row_num > 14:
                    cell.font = Font(size=10)
                
                if col_num == 1 and ':' in str(value):
                    cell.font = Font(bold=True)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40
        
        workbook.save(file_path)
        print(f"✅ Metadata sheet added")


def main():
    """Main execution"""
    
    input_file = r"C:\Users\HP USER\Documents\Data Analyst\pos-eximp-fresh\eximps-cloves quotation.xlsx"
    
    print("="*70)
    print("PROCUREMENT DATA STRUCTURING TOOL")
    print("="*70)
    print()
    
    # Initialize structurer
    structurer = ProcurementDataStructurer(input_file)
    
    # Parse and create structured file
    structurer.parse_quotation()
    output_file = structurer.create_structured_file()
    
    # Add metadata
    structurer.add_metadata_sheet(output_file)
    
    # Display summary
    print("\n📋 STRUCTURED DATA SUMMARY:")
    print("-"*70)
    
    df = pd.read_excel(output_file, sheet_name='Procurement')
    print(f"Total Items: {len(df)}")
    print(f"Categories: {df['Category'].unique().tolist()}")
    print(f"Total Amount: ₦{df['Amount'].sum():,.0f}")
    print(f"\nCategory Breakdown:")
    for category in df['Category'].unique():
        cat_total = df[df['Category'] == category]['Amount'].sum()
        print(f"  • {category}: ₦{cat_total:,.0f}")
    
    print("\n" + "="*70)
    print(f"✅ STRUCTURING COMPLETE")
    print(f"Output file: {output_file}")
    print("="*70)


if __name__ == "__main__":
    main()
