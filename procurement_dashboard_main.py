"""
Complete Procurement Dashboard Integration
===========================================
Main orchestrator for the complete procurement dashboard system.
Handles data structuring, analysis, and report generation.
"""

import sys
import os
from pathlib import Path

# Ensure root is in path
sys.path.append(os.getcwd())

from structure_procurement_data import ProcurementDataStructurer
from procurement_dashboard import ProcurementAnalyzer, run_dashboard_analysis
import pandas as pd
from datetime import datetime


def create_sample_procurement_data():
    """Create a sample structured procurement file for testing"""
    
    data = {
        'Date': [
            '2026-04-22', '2026-04-22', '2026-04-22', '2026-04-22', '2026-04-22',
            '2026-04-22', '2026-04-22', '2026-04-22', '2026-04-22', '2026-04-22',
            '2026-04-22', '2026-04-22', '2026-04-22', '2026-04-22', '2026-04-22',
            '2026-04-22', '2026-04-22'
        ],
        'Item Description': [
            'Cement', 'Blocks', 'Water Container Truck', 'Labour (10 Workers)',
            'Sharp Sand', 'Gravel', 'Iron Rods (12mm)', 'Binding Wire',
            'Transportation', 'Accommodation', 'Miscellaneous',
            'Excavator', 'Pay Loader', 'Hotel Accommodation',
            'Feeding Allowance', 'Lowbed Truck', 'Project Supervision'
        ],
        'Category': [
            'Fencing', 'Fencing', 'Fencing', 'Fencing',
            'Fencing', 'Fencing', 'Fencing', 'Fencing',
            'Fencing', 'Fencing', 'Fencing',
            'Site Clearing', 'Site Clearing', 'Site Clearing',
            'Site Clearing', 'Site Clearing', 'General'
        ],
        'Quantity': [
            300, 15000, 3, 1, 3, 3, 50, 5, 1, 1, 1,
            3, 3, 3, 3, 2, 1
        ],
        'Unit': [
            'Bags', 'Pcs', 'Trips', 'Lot', 'Tons', 'Tons', 'Rods', 'Coils',
            'Lot', 'Lot', 'Lot', 'Days', 'Days', 'Days', 'Days', 'Days', 'Lot'
        ],
        'Unit Price': [
            13000, 600, 15000, 1000000, 3000000, 350000, 10000, 7000,
            100000, 100000, 100000, 450000, 380000, 20000, 10000, 400000, 500000
        ],
        'Amount': [
            3900000, 9000000, 45000, 1000000, 9000000, 1050000, 500000, 35000,
            100000, 100000, 50000, 1350000, 1140000, 60000, 30000, 800000, 500000
        ],
        'Budget': [
            3900000, 9000000, 45000, 1000000, 9000000, 1050000, 500000, 35000,
            100000, 100000, 50000, 1350000, 1140000, 60000, 30000, 800000, 500000
        ],
        'Status': [
            'Pending', 'Pending', 'Approved', 'Approved', 'Pending', 'Pending',
            'Pending', 'Pending', 'Approved', 'Approved', 'Pending',
            'Approved', 'Approved', 'Pending', 'Pending', 'Approved', 'Pending'
        ],
        'Vendor': [
            'Building Supplies Ltd', 'Block Manufacturing Co', 'Water Services',
            'Contractors Association', 'Sand Suppliers', 'Aggregate Suppliers',
            'Steel Products Inc', 'Hardware Store', 'Transport Company',
            'Hotels & Lodging', 'General Supplies',
            'Heavy Equipment Rental', 'Heavy Equipment Rental', 'Hotels & Lodging',
            'Catering Services', 'Transport Company', 'Project Management Ltd'
        ],
        'Payment Method': [
            'Bank Transfer', 'Bank Transfer', 'Cash', 'Bank Transfer',
            'Bank Transfer', 'Bank Transfer', 'Bank Transfer', 'Cash',
            'Bank Transfer', 'Bank Transfer', 'Cash',
            'Bank Transfer', 'Bank Transfer', 'Bank Transfer',
            'Bank Transfer', 'Bank Transfer', 'Bank Transfer'
        ],
        'Notes': [
            'Fencing materials - Agbowa', 'Fencing materials - Agbowa',
            'Water supply for site', 'Construction labor',
            'Foundation base', 'Foundation base', 'Structural reinforcement',
            'Binding material', 'Delivery to site', 'Worker accommodation',
            'Site miscellaneous items', 'Land clearance', 'Site preparation',
            'Supervision accommodation', 'Worker meals', 'Equipment transport',
            'Project oversight and management'
        ]
    }
    
    return pd.DataFrame(data)


def structure_existing_data():
    """Structure existing quotation file"""
    
    print("\n" + "="*80)
    print("STEP 1: STRUCTURING EXISTING QUOTATION DATA")
    print("="*80)
    
    input_file = r"C:\Users\HP USER\Documents\Data Analyst\pos-eximp-fresh\eximps-cloves quotation.xlsx"
    
    try:
        structurer = ProcurementDataStructurer(input_file)
        structurer.parse_quotation()
        output_file = structurer.create_structured_file()
        structurer.add_metadata_sheet(output_file)
        
        print(f"\n✅ Data structured successfully: {output_file}")
        return output_file
    
    except Exception as e:
        print(f"⚠️  Error structuring existing data: {e}")
        print("   Creating sample data instead...")
        return None


def create_sample_file():
    """Create a sample structured procurement file"""
    
    print("\n" + "="*80)
    print("CREATING SAMPLE PROCUREMENT DATA")
    print("="*80)
    
    df = create_sample_procurement_data()
    
    sample_file = r"C:\Users\HP USER\Documents\Data Analyst\pos-eximp-fresh\procurement_data_sample.xlsx"
    
    with pd.ExcelWriter(sample_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Procurement', index=False)
        
        # Format the sheet
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        workbook = writer.book
        worksheet = writer.sheets['Procurement']
        
        # Format header
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # Format data rows
        for row_num in range(2, len(df) + 2):
            for col_num, col in enumerate(df.columns, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.border = border
                
                if col in ['Unit Price', 'Amount', 'Budget']:
                    cell.number_format = '₦#,##0'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                elif col == 'Quantity':
                    cell.number_format = '0'
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Adjust column widths
        widths = {
            'A': 12, 'B': 25, 'C': 15, 'D': 12, 'E': 10,
            'F': 15, 'G': 15, 'H': 15, 'I': 12, 'J': 20,
            'K': 15, 'L': 30
        }
        for col, width in widths.items():
            worksheet.column_dimensions[col].width = width
        
        worksheet.freeze_panes = 'A2'
    
    print(f"✅ Sample file created: {sample_file}")
    return sample_file


def run_full_analysis(data_file):
    """Run comprehensive analysis on procurement data"""
    
    print("\n" + "="*80)
    print("STEP 2: RUNNING COMPREHENSIVE PROCUREMENT ANALYSIS")
    print("="*80)
    print()
    
    try:
        run_dashboard_analysis(data_file)
    except Exception as e:
        print(f"⚠️  Error running analysis: {e}")
        import traceback
        traceback.print_exc()


def generate_ceo_report(data_file):
    """Generate executive CEO report"""
    
    print("\n" + "="*80)
    print("STEP 3: GENERATING EXECUTIVE CEO REPORT")
    print("="*80)
    print()
    
    try:
        analyzer = ProcurementAnalyzer(data_file)
        summary = analyzer.generate_executive_summary()
        
        # Create a nicely formatted CEO report
        report_path = r"C:\Users\HP USER\Documents\Data Analyst\pos-eximp-fresh\CEO_PROCUREMENT_REPORT.txt"
        
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write("╔" + "="*78 + "╗\n")
            f.write("║" + " "*20 + "PROCUREMENT DASHBOARD - CEO EXECUTIVE REPORT" + " "*14 + "║\n")
            f.write("╚" + "="*78 + "╝\n\n")
            
            f.write(f"Report Generated: {summary['report_date']}\n")
            f.write(f"Project: {summary['project']}\n")
            f.write("\n" + "─"*80 + "\n\n")
            
            f.write("💰 FINANCIAL OVERVIEW\n")
            f.write("─"*80 + "\n")
            for key, value in summary['financial_summary'].items():
                f.write(f"  {key.replace('_', ' ').title():.<40} {value:>20}\n")
            
            f.write("\n📋 PROJECT STATUS\n")
            f.write("─"*80 + "\n")
            for key, value in summary['status_breakdown'].items():
                f.write(f"  {key.title():.<40} {value:>20}\n")
            
            f.write("\n🏗️  EXPENDITURE BY CATEGORY\n")
            f.write("─"*80 + "\n")
            for category, value in sorted(summary['category_breakdown'].items(),
                                         key=lambda x: float(x[1].replace('₦', '').replace(',', '')),
                                         reverse=True):
                f.write(f"  {category:.<40} {value:>20}\n")
            
            f.write("\n📊 KEY PERFORMANCE INDICATORS\n")
            f.write("─"*80 + "\n")
            for key, value in summary['key_metrics'].items():
                f.write(f"  {key.replace('_', ' ').title():.<40} {value:>20}\n")
            
            f.write("\n" + "─"*80 + "\n")
            f.write("\n🎯 RECOMMENDATIONS FOR MANAGEMENT\n")
            f.write("─"*80 + "\n")
            
            metrics = analyzer.calculate_metrics()
            
            recommendations = []
            
            if metrics.budget_utilization > 90:
                recommendations.append("  ⚠️  Budget utilization exceeds 90%. Monitor remaining items closely.")
            elif metrics.budget_utilization < 50:
                recommendations.append("  ℹ️  Budget utilization below 50%. Review project timeline and deliverables.")
            
            if metrics.pending_amount > 0:
                recommendations.append(f"  ⏳ {metrics.pending_amount:,.0f} pending approval. Expedite approval process.")
            
            if metrics.cost_variance > 5:
                recommendations.append(f"  📈 Cost variance at {metrics.cost_variance:+.1f}%. Review unit pricing with vendors.")
            elif metrics.cost_variance < -5:
                recommendations.append(f"  📉 Cost variance at {metrics.cost_variance:+.1f}%. Opportunity for cost savings identified.")
            
            if len(recommendations) == 0:
                recommendations.append("  ✅ All metrics within acceptable ranges. Continue current procurement strategy.")
            
            for rec in recommendations:
                f.write(rec + "\n")
            
            f.write("\n" + "╔" + "="*78 + "╗\n")
            f.write("║" + " "*25 + "END OF EXECUTIVE REPORT" + " "*32 + "║\n")
            f.write("╚" + "="*78 + "╝\n")
        
        print(f"✅ CEO Report generated: {report_path}")
        
        # Also display it
        with open(report_path, 'r', encoding='utf-8') as f:
            print(f.read())
    
    except Exception as e:
        print(f"⚠️  Error generating CEO report: {e}")
        import traceback
        traceback.print_exc()


def main():
    """Main orchestration function"""
    
    print("\n")
    print("╔" + "="*78 + "╗")
    print("║" + " "*15 + "PROCUREMENT DASHBOARD - COMPLETE INTEGRATION SYSTEM" + " "*13 + "║")
    print("╚" + "="*78 + "╝")
    
    print("\nThis system provides:")
    print("  ✓ Data structuring and validation")
    print("  ✓ Professional procurement analysis")
    print("  ✓ Executive reporting for CEO/Management")
    print("  ✓ Visual dashboards and charts")
    print("  ✓ Budget tracking and forecasting")
    
    # Try to structure existing data
    structured_file = structure_existing_data()
    
    # If structuring failed, create sample data
    if structured_file is None:
        data_file = create_sample_file()
    else:
        data_file = structured_file
    
    # Run comprehensive analysis
    run_full_analysis(data_file)
    
    # Generate CEO report
    generate_ceo_report(data_file)
    
    print("\n" + "="*80)
    print("✅ COMPLETE PROCUREMENT DASHBOARD SYSTEM READY")
    print("="*80)
    print("\n📁 Generated Files:")
    print(f"  • Structured Data: {data_file}")
    print(f"  • Dashboard Report: procurement_report.xlsx")
    print(f"  • CEO Report: CEO_PROCUREMENT_REPORT.txt")
    print(f"  • Visualizations: procurement_reports/ directory")
    print("\n" + "="*80 + "\n")


if __name__ == "__main__":
    main()
